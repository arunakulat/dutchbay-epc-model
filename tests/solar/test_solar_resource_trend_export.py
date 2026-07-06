"""Long-term solar-GHI trend -> frozen export -> workbook ResourceTrend sheet (#727).

The SOLAR analogue of ``tests/wind/test_resource_trend_export.py`` (#656). It pins:

* ``solar_resource.long_term_trend`` recovers a KNOWN positive (brightening) and negative
  (dimming) trend on a synthetic multi-decade annual-GHI series (Mann-Kendall + Sen's slope);
* ``build_solar_resource_trend_export_block`` encodes the live analysis exactly like
  ``analytics.executive_workbook.serialize_resource_trend`` and DEGRADES EXPLICITLY on a
  short record — never a spurious tau;
* the JSON-safe block is the SAME shape the wind producer emits (``analyzed`` +
  ``summary_records`` of Metric/Value rows) and round-trips through the workbook decoder
  ``resource_trend_df_from_solar_export``;
* ``emit_executive_workbook_from_pipeline`` fills ONE "ResourceTrend" sheet from a solar
  export, wind takes precedence on a hybrid, and absence/opt-out omits the sheet (the
  five-finance-sheet workbook is byte-identical to no-trend).

Data source: the multi-decade annual-GHI series is FROZEN (PVGIS SARAH-3 in production);
these tests build a SYNTHETIC frozen series — no live fetch (#656 / #469 discipline).

Framework: TEST-01 golden pin; CESSPIT fail-loud (explicit short-series degrade);
report/VALIDATE-only (no committed solar CF/AEP or KPI touched).
"""

from __future__ import annotations

import numpy as np
import pandas as pd
import pytest

from analytics.executive_workbook import (
    build_executive_workbook,
    emit_executive_workbook_from_pipeline,
    resource_trend_df_from_solar_export,
    serialize_resource_trend,
)
from solar_resource.long_term_trend import (
    MIN_TREND_YEARS,
    PERIOD_ENERGY_BASIS,
    analyze_long_term_solar_resource,
    build_solar_resource_trend_export_block,
    compute_solar_trend,
    render_trend_markdown,
)

BASE_GHI = 2000.0  # Puttalam-ish annual GHI (kWh/m2/yr)


def _annual_ghi(
    start: int,
    end: int,
    *,
    slope_per_yr: float = 0.0,
    noise: float = 0.0,
    seed: int = 11,
) -> pd.DataFrame:
    """Synthetic FROZEN annual-GHI series (one row per calendar year, year-end index)."""
    years = np.arange(start, end + 1)
    rng = np.random.default_rng(seed)
    vals = (
        BASE_GHI + slope_per_yr * (years - start) + rng.normal(0.0, noise, len(years))
    )
    idx = pd.to_datetime([f"{y}-12-31" for y in years])
    return pd.DataFrame({"ghi_kwh_m2": vals}, index=idx)


# ---------------------------------------------------------------------------
# Trend recovery: known positive (brightening) / negative (dimming)
# ---------------------------------------------------------------------------


class TestTrendRecovery:
    def test_recovers_positive_brightening(self) -> None:
        annual = _annual_ghi(1990, 2024, slope_per_yr=+4.0, noise=5.0)["ghi_kwh_m2"]
        annual.index = annual.index.year
        t = compute_solar_trend(annual)
        assert t.mk_tau > 0
        assert t.sen_slope_per_decade > 0
        assert t.significant is True
        assert "brightening" in t.classification_note

    def test_recovers_negative_dimming(self) -> None:
        annual = _annual_ghi(1990, 2024, slope_per_yr=-4.0, noise=5.0)["ghi_kwh_m2"]
        annual.index = annual.index.year
        t = compute_solar_trend(annual)
        assert t.mk_tau < 0
        assert t.sen_slope_per_decade < 0
        assert t.significant is True
        assert "dimming" in t.classification_note

    def test_flat_series_is_decadal_variability(self) -> None:
        annual = _annual_ghi(1990, 2024, slope_per_yr=0.0, noise=8.0)["ghi_kwh_m2"]
        annual.index = annual.index.year
        t = compute_solar_trend(annual)
        assert t.significant is False
        assert t.classification == "decadal_variability"

    def test_constant_series_no_nan_and_json_safe(self) -> None:
        # LOW #727-review: a zero-variance GHI series must not emit NaN tau/p/r2 that
        # a strict JSON encoder rejects — it degrades to a serializable no-trend.
        import json
        import math

        annual = _annual_ghi(1990, 2024, slope_per_yr=0.0, noise=0.0)["ghi_kwh_m2"]
        annual.index = annual.index.year
        t = compute_solar_trend(annual)
        for v in (t.mk_tau, t.mk_pvalue, t.ols_r2, t.sen_slope_per_decade, t.cov_pct):
            assert math.isfinite(v)
        assert t.significant is False
        assert t.classification == "decadal_variability"
        # The full frozen block round-trips through strict JSON (allow_nan=False).
        block = build_solar_resource_trend_export_block(
            _annual_ghi(1990, 2024, slope_per_yr=0.0, noise=0.0)
        )
        json.dumps(block, allow_nan=False)


# ---------------------------------------------------------------------------
# build_solar_resource_trend_export_block — encoder + explicit degrade
# ---------------------------------------------------------------------------


class TestBuildBlock:
    def test_analyzed_equals_live_serialize(self) -> None:
        series = _annual_ghi(1995, 2024, slope_per_yr=+3.0, noise=6.0)
        block = build_solar_resource_trend_export_block(series)
        assert block["analyzed"] is True
        assert block["summary_records"]
        assert block == serialize_resource_trend(
            analyze_long_term_solar_resource(series)
        )

    def test_block_shape_matches_wind_contract(self) -> None:
        # Same two keys + Metric/Value record shape as the wind producer's block.
        block = build_solar_resource_trend_export_block(
            _annual_ghi(1995, 2024, slope_per_yr=2.0)
        )
        assert set(block) == {"analyzed", "summary_records"}
        for rec in block["summary_records"]:
            assert set(rec) == {"Metric", "Value"}

    def test_energy_base_scales_reference_periods(self) -> None:
        series = _annual_ghi(1995, 2024, slope_per_yr=+3.0, noise=4.0)
        out = analyze_long_term_solar_resource(
            series, base_energy_gwh=464.5, base_cf=0.354
        )
        recs = {r["Metric"]: r["Value"] for r in out["summary_df"].to_dict("records")}
        # GHI-scaled energy + CF surface only when a base is supplied.
        assert any(m.startswith("AEP P50 ") for m in recs)
        assert any(m.startswith("CF ") for m in recs)
        assert recs["Recommended P50 (GWh)"] is not None
        # MEDIUM #727-review: the proxy nature is stated IN the artifact (a "Method" row)
        # and the derived energy is not over-precise (1 dp) — a lender must not read it
        # as a precise pvlib P50.
        assert "proxy" in recs["AEP basis (per-period energy/CF)"].lower()
        assert "not a pvlib re-run" in recs["AEP basis (per-period energy/CF)"]
        aep_val = next(v for m, v in recs.items() if m.startswith("AEP P50 "))
        assert round(aep_val, 1) == aep_val  # <=1 decimal place, no false precision

    def test_markdown_header_caveats_the_ghi_proxy(self) -> None:
        series = _annual_ghi(1995, 2024, slope_per_yr=+3.0, noise=4.0)
        out = analyze_long_term_solar_resource(
            series, base_energy_gwh=464.5, base_cf=0.354
        )
        md = out["markdown"]
        # render_trend_markdown built the same string analyze_* returns.
        assert md == render_trend_markdown(
            out["trend"], out["period_ghi"], out["recommendation"]
        )
        assert PERIOD_ENERGY_BASIS in md
        assert "GHI-scaled AEP by reference period" in md
        # The precise-modelled-sounding bare header is NOT used when energy is present.
        assert "**GHI by reference period:**" not in md
        # The recommended-P50 headline itself carries the proxy caveat.
        assert PERIOD_ENERGY_BASIS in md.split("Recommended P50 basis")[1]

    def test_no_energy_base_omits_energy_rows_and_method(self) -> None:
        out = analyze_long_term_solar_resource(
            _annual_ghi(1995, 2024, slope_per_yr=2.0)
        )
        metrics = [r["Metric"] for r in out["summary_df"].to_dict("records")]
        assert not any(m.startswith("AEP P50 ") for m in metrics)
        assert not any(m.startswith("CF ") for m in metrics)
        # No spurious Method row on a GHI-only sheet.
        assert "AEP basis (per-period energy/CF)" not in metrics
        # The bare (non-proxy) header IS used when there is no derived energy.
        assert "**GHI by reference period:**" in out["markdown"]

    def test_round_trips_into_workbook_decoder(self) -> None:
        series = _annual_ghi(1995, 2024, slope_per_yr=+3.0, noise=4.0)
        block = build_solar_resource_trend_export_block(series)
        got = resource_trend_df_from_solar_export({"long_term_trend": block})
        assert got is not None
        pd.testing.assert_frame_equal(
            got.reset_index(drop=True),
            analyze_long_term_solar_resource(series)["summary_df"].reset_index(
                drop=True
            ),
        )

    def test_nested_under_cashflow_export_is_decoded(self) -> None:
        block = build_solar_resource_trend_export_block(
            _annual_ghi(1995, 2024, slope_per_yr=2.0)
        )
        got = resource_trend_df_from_solar_export(
            {"cashflow_export": {"long_term_trend": block}}
        )
        assert got is not None

    def test_short_series_degrades_explicitly(self) -> None:
        block = build_solar_resource_trend_export_block(_annual_ghi(2022, 2024))
        assert block["analyzed"] is False
        assert "summary_records" not in block
        assert "3 calendar year" in block["reason"]
        assert "10-yr minimum" in block["reason"]

    def test_min_years_boundary(self) -> None:
        end = 2024
        assert (
            build_solar_resource_trend_export_block(
                _annual_ghi(end - MIN_TREND_YEARS + 1, end)
            )["analyzed"]
            is True
        )
        assert (
            build_solar_resource_trend_export_block(
                _annual_ghi(end - MIN_TREND_YEARS + 2, end)
            )["analyzed"]
            is False
        )


# ---------------------------------------------------------------------------
# Workbook consumer: fills / omits the ResourceTrend sheet
# ---------------------------------------------------------------------------


@pytest.fixture()
def pipeline_result() -> dict:
    return {
        "kpis": {
            "scenario_name": "solar_case",
            "project_irr": 0.02,
            "min_dscr": 1.3,
        },
        "annual_rows": [{"year": 1, "cfads": 10.0}],
        "debt_result": {},
        "scenario_result": {"debt_covenants": {}},
    }


_FINANCE_SHEETS = ["Summary", "Cashflow", "DebtService", "Ratios", "ScenarioSummary"]


class TestWorkbookConsumer:
    def test_solar_export_fills_resource_trend(self, pipeline_result, tmp_path) -> None:
        import openpyxl

        block = build_solar_resource_trend_export_block(
            _annual_ghi(1995, 2024, slope_per_yr=+3.0, noise=4.0)
        )
        solar_export = {
            "cashflow_export": {"scenario": "P50"},
            "long_term_trend": block,
        }
        out = emit_executive_workbook_from_pipeline(
            pipeline_result, tmp_path / "wb.xlsx", solar_export=solar_export
        )
        names = openpyxl.load_workbook(out).sheetnames
        assert "ResourceTrend" in names
        wb = openpyxl.load_workbook(out)
        header = [c.value for c in wb["ResourceTrend"][1]]
        assert header == ["Metric", "Value"]
        # The solar-signed "Resource" row: Metric "Resource", Value "Solar GHI".
        values = [row[1].value for row in wb["ResourceTrend"].iter_rows(min_row=2)]
        assert "Solar GHI" in values

    def test_absent_solar_omits_sheet_byte_identical(
        self, pipeline_result, tmp_path
    ) -> None:
        import openpyxl

        a = emit_executive_workbook_from_pipeline(pipeline_result, tmp_path / "a.xlsx")
        b = emit_executive_workbook_from_pipeline(
            pipeline_result, tmp_path / "b.xlsx", solar_export={"other": 1}
        )
        assert openpyxl.load_workbook(a).sheetnames == _FINANCE_SHEETS
        assert openpyxl.load_workbook(b).sheetnames == _FINANCE_SHEETS
        assert (tmp_path / "a.xlsx").read_bytes() == (tmp_path / "b.xlsx").read_bytes()

    def test_optout_degraded_block_omits_sheet(self, pipeline_result, tmp_path) -> None:
        import openpyxl

        out = emit_executive_workbook_from_pipeline(
            pipeline_result,
            tmp_path / "wb.xlsx",
            solar_export={"long_term_trend": {"analyzed": False, "reason": "short"}},
        )
        assert openpyxl.load_workbook(out).sheetnames == _FINANCE_SHEETS

    def test_wind_takes_precedence_on_hybrid(self, pipeline_result, tmp_path) -> None:
        # Both a wind and a solar trend present → the single sheet is WIND's.
        import openpyxl

        solar_block = build_solar_resource_trend_export_block(
            _annual_ghi(1995, 2024, slope_per_yr=+3.0)
        )
        # A minimal hand-built wind block (same contract) marked with a wind signature.
        wind_df = pd.DataFrame(
            [
                {"Metric": "Resource", "Value": "Wind"},
                {"Metric": "Mean wind speed", "Value": 8.0},
            ]
        )
        wind_block = serialize_resource_trend({"summary_df": wind_df})
        out = emit_executive_workbook_from_pipeline(
            pipeline_result,
            tmp_path / "wb.xlsx",
            wind_export={"long_term_trend": wind_block},
            solar_export={"long_term_trend": solar_block},
        )
        wb = openpyxl.load_workbook(out)
        assert "ResourceTrend" in wb.sheetnames
        values = [row[1].value for row in wb["ResourceTrend"].iter_rows(min_row=2)]
        assert "Wind" in values
        assert "Solar GHI" not in values

    def test_build_executive_workbook_writes_solar_sheet(self, tmp_path) -> None:
        # The raw builder writes whatever resource_trend_df it is handed (tech-agnostic).
        import openpyxl

        df = analyze_long_term_solar_resource(
            _annual_ghi(1995, 2024, slope_per_yr=2.0)
        )["summary_df"]
        empty = pd.DataFrame({"x": [1, 2, 3]})
        out = build_executive_workbook(
            empty,
            empty,
            empty,
            empty,
            empty,
            tmp_path / "wb.xlsx",
            resource_trend_df=df,
        )
        assert "ResourceTrend" in openpyxl.load_workbook(out).sheetnames
