#!/usr/bin/env python3
"""Coverage tests for analytics.export_helpers.

The module is a pure DataFrame -> (Excel | matplotlib chart) export layer with
no network or config dependencies, so these tests construct small in-memory
DataFrames and assert the produced artifacts:

- validate_for_export: every documented warning branch
- ExcelExporter: cover/metadata sheet, generic dataframe sheet (headers,
  auto-filter, freeze panes, warning highlight), conditional formatting,
  chart-image embedding, summary/timeseries writers, the high-level
  export_summary_and_timeseries entry point, board views, sheet reordering,
  autofit, and the no-writer early returns.
- ChartExporter / ChartGenerator: every public plot method, both the happy
  path (file written) and the matplotlib-unavailable / missing-column / empty
  paths, plus the ValueError raised when no numeric KPI column exists.

matplotlib is forced onto the non-interactive Agg backend at import time so no
display is required. All workbooks/images are written under pytest's tmp_path.
"""

from __future__ import annotations

from pathlib import Path
from typing import Any

import matplotlib
import pandas as pd
import pytest
from openpyxl import load_workbook

matplotlib.use("Agg")  # headless; must precede any pyplot import

from analytics.export_helpers import (  # noqa: E402  (after backend selection)
    ChartExporter,
    ChartGenerator,
    ExcelExporter,
    validate_for_export,
)

REPO_ROOT = Path(__file__).resolve().parents[2]


# ---------------------------------------------------------------------------
# DataFrame fixtures
# ---------------------------------------------------------------------------
@pytest.fixture
def summary_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "scenario_name": ["base", "stress"],
            "project_irr": [0.07, 0.04],
            "equity_irr": [0.05, 0.01],
        }
    )


@pytest.fixture
def timeseries_df() -> pd.DataFrame:
    return pd.DataFrame(
        {
            "scenario_name": ["base", "base", "stress", "stress"],
            "year": [1, 2, 1, 2],
            "dscr": [1.45, 1.10, 1.30, 0.95],
        }
    )


# ---------------------------------------------------------------------------
# validate_for_export
# ---------------------------------------------------------------------------
def test_validate_for_export_no_warnings(
    summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    assert validate_for_export(summary_df, timeseries_df) == []


def test_validate_for_export_all_warnings() -> None:
    empty = pd.DataFrame({"unrelated": [1]})
    warnings = validate_for_export(empty, empty)
    joined = " ".join(warnings)
    assert "summary_df missing 'scenario_name' column" in warnings
    assert "timeseries_df missing 'scenario_name' column" in warnings
    assert "project_irr" in joined
    assert "equity_irr" in joined
    assert "summary_df missing expected KPI column 'project_irr'" in warnings
    assert "timeseries_df missing 'dscr' column" in warnings


# ---------------------------------------------------------------------------
# ExcelExporter: save() no-op when never opened
# ---------------------------------------------------------------------------
def test_save_without_writer_is_noop(tmp_path: Path) -> None:
    exporter = ExcelExporter(tmp_path / "empty.xlsx")
    # No sheets created -> writer is None -> save() must not raise or write.
    exporter.save()
    assert exporter._writer is None


# ---------------------------------------------------------------------------
# add_metadata_sheet
# ---------------------------------------------------------------------------
def test_add_metadata_sheet_with_run_id(tmp_path: Path) -> None:
    out = tmp_path / "meta.xlsx"
    exporter = ExcelExporter(out)
    exporter.add_metadata_sheet(
        metadata={"Project": "DutchBay", "MW": 150},
        run_id="run-001",
    )
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Report_Cover"]
    assert ws["A1"].value == "Field"
    assert ws["B1"].value == "Value"
    # Header is bolded (best-effort styling path executed).
    assert ws["A1"].font.bold is True
    rows = {row[0]: row[1] for row in ws.iter_rows(min_row=2, values_only=True)}
    assert "Report Generated" in rows
    assert rows["Run ID"] == "run-001"
    assert rows["Project"] == "DutchBay"
    assert rows["MW"] == "150"
    # Column widths set by the styling block.
    assert ws.column_dimensions["A"].width == 30
    assert ws.column_dimensions["B"].width == 50


def test_add_metadata_sheet_without_run_id(tmp_path: Path) -> None:
    out = tmp_path / "meta2.xlsx"
    exporter = ExcelExporter(out)
    exporter.add_metadata_sheet(metadata={"k": "v"})
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Report_Cover"]
    fields = [row[0] for row in ws.iter_rows(min_row=2, values_only=True)]
    assert "Run ID" not in fields
    assert "Report Generated" in fields


# ---------------------------------------------------------------------------
# add_dataframe_sheet: header formatting, auto-filter, freeze panes, highlight
# ---------------------------------------------------------------------------
def test_add_dataframe_sheet_full_options(tmp_path: Path) -> None:
    out = tmp_path / "df.xlsx"
    exporter = ExcelExporter(out)
    df = pd.DataFrame(
        {
            "scenario_name": ["a", "b", "c"],
            "dscr": [1.5, 1.0, "n/a"],  # non-numeric value exercises except branch
        }
    )
    exporter.add_dataframe_sheet(
        sheet_name="Data",
        df=df,
        freeze_panes="A2",
        format_headers=True,
        auto_filter=True,
        highlight_warnings=True,
        warning_column="dscr",
        warning_threshold=1.2,
    )
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Data"]
    assert ws["A1"].value == "scenario_name"
    assert ws["A1"].font.bold is True
    assert ws.freeze_panes == "A2"
    assert ws.auto_filter.ref == ws.dimensions
    # dscr=1.0 (< 1.2) on the "b" row (row 3) is highlighted; 1.5 is not.
    assert ws["B3"].fill.start_color.rgb.endswith("FFCCCC")
    assert ws["B2"].fill.fill_type in (None, "none")


def test_add_dataframe_sheet_minimal_no_options(tmp_path: Path) -> None:
    out = tmp_path / "df_min.xlsx"
    exporter = ExcelExporter(out)
    df = pd.DataFrame({"x": [1, 2]})
    exporter.add_dataframe_sheet(
        sheet_name="Plain",
        df=df,
        format_headers=False,
        auto_filter=False,
    )
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Plain"]
    assert ws["A1"].value == "x"
    # format_headers / auto_filter disabled: no auto-filter range registered.
    assert ws.auto_filter.ref is None


def test_add_dataframe_sheet_highlight_missing_column(tmp_path: Path) -> None:
    # warning_column not in df -> highlight block runs but the inner guard skips.
    out = tmp_path / "df_miss.xlsx"
    exporter = ExcelExporter(out)
    df = pd.DataFrame({"a": [1, 2]})
    exporter.add_dataframe_sheet(
        sheet_name="Miss",
        df=df,
        highlight_warnings=True,
        warning_column="not_here",
        warning_threshold=1.0,
    )
    exporter.save()
    wb = load_workbook(out)
    assert "Miss" in wb.sheetnames


# ---------------------------------------------------------------------------
# add_conditional_formatting
# ---------------------------------------------------------------------------
def test_add_conditional_formatting_applies_rule(tmp_path: Path) -> None:
    out = tmp_path / "cf.xlsx"
    exporter = ExcelExporter(out)
    exporter.add_dataframe_sheet(
        sheet_name="Summary",
        df=pd.DataFrame({"scenario_name": ["a", "b"], "dscr": [1.4, 0.9]}),
    )
    exporter.add_conditional_formatting(
        sheet_name="Summary",
        column_range="B2:B3",
        rule_type="lessThan",
        threshold=1.2,
    )
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Summary"]
    # Conditional formatting rule registered for the requested range.
    assert any("B2:B3" in str(rng) for rng in ws.conditional_formatting)


def test_add_conditional_formatting_none_threshold_skips(tmp_path: Path) -> None:
    out = tmp_path / "cf_none.xlsx"
    exporter = ExcelExporter(out)
    exporter.add_dataframe_sheet(
        sheet_name="Summary",
        df=pd.DataFrame({"scenario_name": ["a"], "dscr": [1.0]}),
    )
    exporter.add_conditional_formatting(
        sheet_name="Summary",
        column_range="B2:B2",
        threshold=None,
    )
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Summary"]
    # No rule added because threshold was None.
    assert len(list(ws.conditional_formatting)) == 0


# ---------------------------------------------------------------------------
# add_chart_image
# ---------------------------------------------------------------------------
def _write_png(path: Path) -> None:
    import matplotlib.pyplot as plt

    fig, ax = plt.subplots()
    ax.plot([0, 1], [0, 1])
    fig.savefig(path)
    plt.close(fig)


def test_add_chart_image_embeds(tmp_path: Path) -> None:
    img = tmp_path / "chart.png"
    _write_png(img)

    out = tmp_path / "with_img.xlsx"
    exporter = ExcelExporter(out)
    exporter.add_dataframe_sheet(
        sheet_name="Summary",
        df=pd.DataFrame({"a": [1]}),
    )
    exporter.add_chart_image(sheet_name="Summary", image_path=str(img), cell="D2")
    exporter.save()

    wb = load_workbook(out)
    ws = wb["Summary"]
    assert len(ws._images) == 1


def test_add_chart_image_bad_path_is_nonfatal(tmp_path: Path) -> None:
    out = tmp_path / "bad_img.xlsx"
    exporter = ExcelExporter(out)
    exporter.add_dataframe_sheet(
        sheet_name="Summary",
        df=pd.DataFrame({"a": [1]}),
    )
    # Nonexistent image path -> XLImage raises, caught and logged, no exception.
    exporter.add_chart_image(
        sheet_name="Summary",
        image_path=str(tmp_path / "missing.png"),
    )
    exporter.save()
    wb = load_workbook(out)
    assert len(wb["Summary"]._images) == 0


# ---------------------------------------------------------------------------
# write_scenario_summary / write_scenario_timeseries
# ---------------------------------------------------------------------------
def test_write_scenario_summary_and_timeseries(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    out = tmp_path / "sheets.xlsx"
    exporter = ExcelExporter(out)
    exporter.write_scenario_summary(summary_df)
    exporter.write_scenario_timeseries(timeseries_df)
    exporter.save()

    wb = load_workbook(out)
    assert "Summary" in wb.sheetnames
    assert "Timeseries" in wb.sheetnames


# ---------------------------------------------------------------------------
# export_summary_and_timeseries: the high-level entry point
# ---------------------------------------------------------------------------
def test_export_summary_and_timeseries_full(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    out = tmp_path / "full.xlsx"
    exporter = ExcelExporter(out)
    exporter.export_summary_and_timeseries(
        summary_df=summary_df,
        timeseries_df=timeseries_df,
        add_board_views=True,
        scenario_metadata={"Project": "DutchBay"},
        run_id="run-xyz",
        validate_before_export=True,
    )

    wb = load_workbook(out)
    names = wb.sheetnames
    # Board reordering puts cover/summary first, timeseries after the views.
    assert names[0] == "Report_Cover"
    assert names[1] == "Summary"
    assert "DSCR_View" in names
    assert "IRR_View" in names
    assert names.index("Timeseries") > names.index("IRR_View")


def test_export_summary_and_timeseries_minimal(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    out = tmp_path / "minimal.xlsx"
    exporter = ExcelExporter(out)
    exporter.export_summary_and_timeseries(
        summary_df=summary_df,
        timeseries_df=timeseries_df,
        add_board_views=False,
        scenario_metadata=None,
        validate_before_export=False,
    )

    wb = load_workbook(out)
    names = wb.sheetnames
    assert "Report_Cover" not in names
    assert "DSCR_View" not in names
    assert "Summary" in names
    assert "Timeseries" in names


# ---------------------------------------------------------------------------
# export_summary_and_timeseries: board-deck enrichments (#662)
# ---------------------------------------------------------------------------
def test_export_dscr_conditional_formatting_and_embedded_charts(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    """The opt-in #662 knobs add a live DSCR rule + a Charts sheet with images."""
    img = tmp_path / "board.png"
    _write_png(img)

    out = tmp_path / "board_deck.xlsx"
    exporter = ExcelExporter(out)
    exporter.export_summary_and_timeseries(
        summary_df=summary_df,
        timeseries_df=timeseries_df,
        add_board_views=True,
        dscr_conditional_threshold=1.2,
        embed_chart_images=[str(img)],
    )

    wb = load_workbook(out)
    # DSCR_View carries a live conditional-formatting rule (not just a static fill).
    dscr_ws = wb["DSCR_View"]
    rules = list(dscr_ws.conditional_formatting)
    assert len(rules) == 1
    # dscr is the 3rd column (scenario_name, Year, dscr) -> column C.
    assert any("C2:C" in str(r) for r in rules)
    # Charts sheet exists with the embedded image.
    assert "Charts" in wb.sheetnames
    assert len(wb["Charts"]._images) == 1


def test_export_enrichments_default_off_is_unchanged(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    """With both #662 knobs left at their None default, no rule/Charts sheet appears."""
    out = tmp_path / "no_enrich.xlsx"
    exporter = ExcelExporter(out)
    exporter.export_summary_and_timeseries(
        summary_df=summary_df,
        timeseries_df=timeseries_df,
        add_board_views=True,
    )

    wb = load_workbook(out)
    assert "Charts" not in wb.sheetnames
    assert len(list(wb["DSCR_View"].conditional_formatting)) == 0


def test_embed_chart_images_skips_missing_files(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    """A list of only-missing image paths produces no Charts sheet (nothing to embed)."""
    out = tmp_path / "no_files.xlsx"
    exporter = ExcelExporter(out)
    exporter.export_summary_and_timeseries(
        summary_df=summary_df,
        timeseries_df=timeseries_df,
        add_board_views=True,
        embed_chart_images=[str(tmp_path / "does_not_exist.png")],
    )

    wb = load_workbook(out)
    assert "Charts" not in wb.sheetnames


def test_dscr_conditional_formatting_noop_without_dscr_view(
    tmp_path: Path, summary_df: pd.DataFrame
) -> None:
    """A threshold with no DSCR column in the timeseries adds no rule (no DSCR_View)."""
    ts_no_dscr = pd.DataFrame({"scenario_name": ["base"], "year": [1]})
    out = tmp_path / "no_dscr.xlsx"
    exporter = ExcelExporter(out)
    exporter.export_summary_and_timeseries(
        summary_df=summary_df,
        timeseries_df=ts_no_dscr,
        add_board_views=True,
        dscr_conditional_threshold=1.2,
    )

    wb = load_workbook(out)
    # No DSCR_View sheet was produced, so the conditional-formatting path is a no-op.
    assert "DSCR_View" not in wb.sheetnames


# ---------------------------------------------------------------------------
# _add_board_friendly_views: period-named DSCR + no-IRR summary branches
# ---------------------------------------------------------------------------
def test_board_views_period_column_and_no_irr(tmp_path: Path) -> None:
    out = tmp_path / "period.xlsx"
    exporter = ExcelExporter(out)
    summary = pd.DataFrame({"scenario_name": ["a"], "npv": [10.0]})  # no IRR cols
    timeseries = pd.DataFrame(
        {
            "scenario_name": ["a", "a"],
            "period": [1, 2],
            "dscr": [1.3, 1.1],
        }
    )
    exporter.export_summary_and_timeseries(
        summary_df=summary,
        timeseries_df=timeseries,
        add_board_views=True,
    )

    wb = load_workbook(out)
    assert "DSCR_View" in wb.sheetnames
    # period -> "Period" rename applied.
    assert wb["DSCR_View"]["B1"].value == "Period"
    # No IRR columns in summary -> no IRR_View sheet.
    assert "IRR_View" not in wb.sheetnames


# ---------------------------------------------------------------------------
# _reorder_sheets_for_board / autofit_all: no-writer early returns
# ---------------------------------------------------------------------------
def test_reorder_and_autofit_without_writer_are_noops(tmp_path: Path) -> None:
    exporter = ExcelExporter(tmp_path / "noop.xlsx")
    # Both guard on self._writer is None and must return without raising.
    exporter._reorder_sheets_for_board()
    exporter.autofit_all()
    assert exporter._writer is None


def test_reorder_keeps_non_priority_sheet_last(
    tmp_path: Path, summary_df: pd.DataFrame
) -> None:
    out = tmp_path / "extra.xlsx"
    exporter = ExcelExporter(out)
    # A sheet name that is NOT in the board priority list exercises the
    # "append remaining sheets" tail of _reorder_sheets_for_board.
    exporter.add_dataframe_sheet(sheet_name="Custom_Notes", df=summary_df)
    exporter.write_scenario_summary(summary_df)
    exporter._reorder_sheets_for_board()
    exporter.autofit_all()
    exporter.save()

    wb = load_workbook(out)
    names = wb.sheetnames
    # Summary is a priority sheet (ordered first); the non-priority custom
    # sheet is retained and pushed to the tail.
    assert names[0] == "Summary"
    assert "Custom_Notes" in names
    assert names.index("Custom_Notes") > names.index("Summary")


def test_autofit_handles_empty_worksheet_column(
    tmp_path: Path, summary_df: pd.DataFrame
) -> None:
    out = tmp_path / "autofit_empty.xlsx"
    exporter = ExcelExporter(out)
    exporter.write_scenario_summary(summary_df)
    # Inject a worksheet with no cells: ws.columns then yields an empty
    # column tuple, so max(...) over the empty generator raises ValueError
    # and autofit_all must swallow it (continue) rather than crash.
    writer = exporter._ensure_writer()
    writer.book.create_sheet("Blank")
    exporter.autofit_all()
    exporter.save()

    wb = load_workbook(out)
    assert "Blank" in wb.sheetnames


# ---------------------------------------------------------------------------
# ChartExporter
# ---------------------------------------------------------------------------
def test_chart_exporter_creates_output_dir(tmp_path: Path) -> None:
    target = tmp_path / "nested" / "charts"
    exporter = ChartExporter(target)
    assert target.is_dir()
    assert exporter.output_dir == target


def test_export_dscr_chart_writes_png(
    tmp_path: Path, timeseries_df: pd.DataFrame
) -> None:
    exporter = ChartExporter(tmp_path)
    path = exporter.export_dscr_chart(timeseries_df)
    assert path is not None
    assert path.exists()
    assert path.name == "dscr_series.png"
    assert path.stat().st_size > 0


def test_export_dscr_chart_missing_columns(tmp_path: Path) -> None:
    exporter = ChartExporter(tmp_path)
    df = pd.DataFrame({"scenario_name": ["a"]})  # no dscr column
    assert exporter.export_dscr_chart(df) is None


def test_export_irr_histogram_writes_png(
    tmp_path: Path, summary_df: pd.DataFrame
) -> None:
    exporter = ChartExporter(tmp_path)
    path = exporter.export_irr_histogram(summary_df)
    assert path is not None
    assert path.exists()
    assert path.name == "project_irr_hist.png"


def test_export_irr_histogram_missing_column(tmp_path: Path) -> None:
    exporter = ChartExporter(tmp_path)
    df = pd.DataFrame({"scenario_name": ["a"]})  # no project_irr column
    assert exporter.export_irr_histogram(df) is None


def test_export_irr_histogram_all_nan(tmp_path: Path) -> None:
    exporter = ChartExporter(tmp_path)
    df = pd.DataFrame({"project_irr": ["x", "y"]})  # coerce -> all NaN -> empty
    assert exporter.export_irr_histogram(df) is None


def test_export_charts_dict(
    tmp_path: Path, summary_df: pd.DataFrame, timeseries_df: pd.DataFrame
) -> None:
    exporter = ChartExporter(tmp_path)
    result = exporter.export_charts(summary_df, timeseries_df)
    assert set(result) == {"dscr_chart", "irr_histogram"}
    assert result["dscr_chart"] is not None
    assert result["irr_histogram"] is not None
    assert result["dscr_chart"].exists()


def test_chart_exporter_handles_no_matplotlib(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
    summary_df: pd.DataFrame,
    timeseries_df: pd.DataFrame,
) -> None:
    exporter = ChartExporter(tmp_path)
    monkeypatch.setattr(exporter, "_get_plt", lambda: None)
    assert exporter.export_dscr_chart(timeseries_df) is None
    assert exporter.export_irr_histogram(summary_df) is None


# ---------------------------------------------------------------------------
# ChartGenerator
# ---------------------------------------------------------------------------
def test_chart_generator_creates_output_dir(tmp_path: Path) -> None:
    target = tmp_path / "gen"
    gen = ChartGenerator(target)
    assert target.is_dir()
    assert gen.output_dir == target


def test_plot_kpi_comparison_dataframe_named_column(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    df = pd.DataFrame({"scenario_name": ["a", "b"], "project_irr": [0.07, 0.05]})
    path = gen.plot_kpi_comparison(df, "project_irr", "kpi.png")
    assert path == tmp_path / "kpi.png"
    assert path.exists()


def test_plot_kpi_comparison_mapping_uses_index_labels(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    # Mapping input + the requested KPI IS present, but there is no
    # scenario_name column, so labels fall back to the DataFrame index.
    data: dict[str, list[float]] = {"value": [1.0, 2.0, 3.0]}
    path = gen.plot_kpi_comparison(data, "value", "kpi2.png")
    assert path.exists()


def test_plot_kpi_comparison_missing_kpi_hits_numeric_fallback(
    tmp_path: Path,
) -> None:
    gen = ChartGenerator(tmp_path)
    # Requested KPI absent but a numeric column exists -> the fallback selects
    # the first numeric column and still produces a chart. (Regression: the old
    # `if not numeric_cols` evaluated a pandas Index in boolean context and
    # raised "truth value of an Index is ambiguous" before the fallback ran.)
    df = pd.DataFrame({"label": ["a", "b"], "value": [1.0, 2.0]})
    path = gen.plot_kpi_comparison(df, "absent_kpi", "kpi3.png")
    assert path.exists()


def test_plot_kpi_comparison_no_numeric_columns_raises(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    # No numeric column at all -> the documented "No numeric columns" raise,
    # now reachable since the Index-truthiness bug is fixed.
    df = pd.DataFrame({"label": ["a", "b"], "tag": ["x", "y"]})
    with pytest.raises(ValueError, match="No numeric columns available for KPI chart"):
        gen.plot_kpi_comparison(df, "absent_kpi", "kpi_nonum.png")


def test_plot_kpi_comparison_absolute_path(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path / "gen_dir")
    abs_target = tmp_path / "explicit" / "abs_kpi.png"
    df = pd.DataFrame({"scenario_name": ["a"], "irr": [0.1]})
    path = gen.plot_kpi_comparison(df, "irr", abs_target)
    # Absolute output_file is honoured verbatim (parent created on demand).
    assert path == abs_target
    assert path.exists()


def test_plot_npv_distribution(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    path = gen.plot_npv_distribution([1.0, 2.0, 3.0, 4.0], "npv.png", bins=5)
    assert path.exists()


def test_plot_dscr_comparison_with_threshold(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    data = {"base": [1.4, 1.2, 1.1], "stress": [1.1, 0.9, 0.8]}
    path = gen.plot_dscr_comparison(data, "dscr_cmp.png", threshold=1.0)
    assert path.exists()


def test_plot_dscr_comparison_without_threshold(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    data = {"base": [1.4, 1.2]}
    path = gen.plot_dscr_comparison(data, "dscr_cmp2.png")
    assert path.exists()


def test_plot_debt_waterfall(tmp_path: Path) -> None:
    gen = ChartGenerator(tmp_path)
    # One non-empty series, one empty series -> exercises the s[-1]/0.0 branch.
    data: dict[str, list[float]] = {"base": [10.0, 8.0, 5.0], "empty": []}
    path = gen.plot_debt_waterfall(data, "waterfall.png")
    assert path.exists()


def test_chart_generator_handles_no_matplotlib(
    tmp_path: Path, monkeypatch: pytest.MonkeyPatch
) -> None:
    gen = ChartGenerator(tmp_path)
    monkeypatch.setattr(gen, "_get_plt", lambda: None)

    df = pd.DataFrame({"scenario_name": ["a"], "irr": [0.1]})
    # Each method returns the resolved path without writing a file.
    kpi_path = gen.plot_kpi_comparison(df, "irr", "no_mpl_kpi.png")
    npv_path = gen.plot_npv_distribution([1.0, 2.0], "no_mpl_npv.png")
    dscr_path = gen.plot_dscr_comparison({"a": [1.0]}, "no_mpl_dscr.png")
    debt_path = gen.plot_debt_waterfall({"a": [1.0]}, "no_mpl_debt.png")

    for p in (kpi_path, npv_path, dscr_path, debt_path):
        assert isinstance(p, Path)
        assert not p.exists()


def test_get_plt_returns_module(tmp_path: Path) -> None:
    # Direct sanity check that the matplotlib import path returns a usable module.
    chart_exporter = ChartExporter(tmp_path)
    gen = ChartGenerator(tmp_path)
    plt_a: Any = chart_exporter._get_plt()
    plt_b: Any = gen._get_plt()
    assert plt_a is not None
    assert plt_b is not None
    assert hasattr(plt_a, "figure")
