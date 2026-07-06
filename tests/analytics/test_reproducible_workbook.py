"""Byte-reproducible openpyxl workbook artifacts (#886).

Pins the shared freeze helpers that make ``.xlsx`` emitters deterministic under the
KPI-neutral byte-identity discipline:

* ``normalize_xlsx_reproducible`` removes the two wall-clock sources openpyxl bakes in
  (the ``docProps/core.xml`` ``dcterms:modified`` re-stamp in ``save_workbook`` and the
  per-member ZIP ``date_time``), so two builds of the same data moments apart are
  raw-ZIP-byte identical — the property ``tests/solar/test_solar_resource_trend_export``
  ``::test_absent_solar_omits_sheet_byte_identical`` intermittently failed on (the CI
  flake #886);
* it is idempotent, preserves cell content and sheet structure, and freezes the document
  timestamps to the fixed epoch;
* ``freeze_workbook_reproducible`` freezes an in-memory ``Workbook``'s ``created`` /
  ``modified`` (the object-level companion the grid emitter #884 will reuse).

Framework: TEST-01 golden pin; VALIDATE-only (no finance value touched).
"""

from __future__ import annotations

import io
import time
import zipfile

import openpyxl
import pandas as pd

from analytics.reproducible_workbook import (
    REPRODUCIBLE_EPOCH,
    freeze_workbook_reproducible,
    normalize_xlsx_reproducible,
)


def _build_workbook(path) -> None:
    """Write a small multi-sheet xlsx exactly like the executive-workbook path."""
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame({"Metric": ["a", "b"], "Value": [1, 2]}).to_excel(
            writer, sheet_name="Summary", index=False
        )
        pd.DataFrame({"year": [1, 2], "cfads": [10.0, 11.0]}).to_excel(
            writer, sheet_name="Cashflow", index=False
        )


def test_two_builds_are_byte_identical_across_a_wall_clock_gap(tmp_path) -> None:
    # A >1s gap straddles the second-resolution ZIP date_time and the core.xml
    # modified stamp — exactly the condition that made the byte-identity flaky.
    a = tmp_path / "a.xlsx"
    b = tmp_path / "b.xlsx"
    _build_workbook(a)
    normalize_xlsx_reproducible(a)
    time.sleep(1.1)
    _build_workbook(b)
    normalize_xlsx_reproducible(b)
    assert a.read_bytes() == b.read_bytes()


def test_normalize_is_idempotent(tmp_path) -> None:
    p = tmp_path / "wb.xlsx"
    _build_workbook(p)
    normalize_xlsx_reproducible(p)
    once = p.read_bytes()
    normalize_xlsx_reproducible(p)
    assert p.read_bytes() == once


def test_normalize_preserves_content_and_freezes_timestamps(tmp_path) -> None:
    p = tmp_path / "wb.xlsx"
    _build_workbook(p)
    normalize_xlsx_reproducible(p)
    wb = openpyxl.load_workbook(p)
    assert wb.sheetnames == ["Summary", "Cashflow"]
    assert wb["Summary"]["A1"].value == "Metric"
    assert wb["Cashflow"]["B2"].value == 10.0
    assert wb.properties.created == REPRODUCIBLE_EPOCH
    assert wb.properties.modified == REPRODUCIBLE_EPOCH


def test_normalize_pins_every_zip_member_date_time(tmp_path) -> None:
    p = tmp_path / "wb.xlsx"
    _build_workbook(p)
    normalize_xlsx_reproducible(p)
    with zipfile.ZipFile(io.BytesIO(p.read_bytes())) as zf:
        stamps = {info.date_time for info in zf.infolist()}
    assert stamps == {(1980, 1, 1, 0, 0, 0)}


def test_freeze_workbook_sets_object_level_timestamps() -> None:
    wb = openpyxl.Workbook()
    returned = freeze_workbook_reproducible(wb)
    assert returned is wb  # returns wb for chaining
    assert wb.properties.created == REPRODUCIBLE_EPOCH
    assert wb.properties.modified == REPRODUCIBLE_EPOCH
