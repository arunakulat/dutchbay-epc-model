"""
Smoke tests for analytics.executive_workbook.

These tests deliberately stay high-level:
    - Do we produce an XLSX file?
    - Does it contain the expected sheet names?
    - Do basic shapes match expectations?

The detailed financial correctness is exercised elsewhere in the pipeline.
"""

from pathlib import Path

import pandas as pd
from openpyxl import load_workbook

from analytics.executive_workbook import build_executive_workbook


def _dummy_summary_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "scenario_name": "dutchbay_lendercase_2025Q4",
                "project_irr": 0.125,
                "equity_irr": 0.165,
                "project_npv_lkr": 1_234_567_890.0,
                "equity_npv_lkr": 789_000_000.0,
                "wacc_pct": 0.10,
                "leverage_pct": 0.70,
                "debt_tenor_years": 15,
                "p50_generation_gwh": 650.0,
                "p90_generation_gwh": 580.0,
                "tariff_lkr_per_kwh": 20.30,
                "dscr_min": 1.35,
                "dscr_avg": 1.50,
                "llcr": 1.55,
                "plcr": 1.70,
                "dsra_months": 6.0,
            }
        ]
    )


def _dummy_cashflow_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "year": 0,
                "revenue_lkr": 0.0,
                "opex_lkr": 0.0,
                "cfads_lkr": 0.0,
            },
            {
                "year": 1,
                "revenue_lkr": 1_000_000.0,
                "opex_lkr": 200_000.0,
                "cfads_lkr": 800_000.0,
            },
        ]
    )


def _dummy_debt_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "year": 1,
                "opening_balance_lkr": 10_000_000.0,
                "interest_lkr": 500_000.0,
                "principal_lkr": 700_000.0,
                "debt_service_lkr": 1_200_000.0,
                "cfads_lkr": 1_600_000.0,
                "dscr": 1.33,
            },
            {
                "year": 2,
                "opening_balance_lkr": 9_300_000.0,
                "interest_lkr": 465_000.0,
                "principal_lkr": 735_000.0,
                "debt_service_lkr": 1_200_000.0,
                "cfads_lkr": 1_700_000.0,
                "dscr": 1.42,
            },
        ]
    )


def _dummy_ratios_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "metric_name": "min_dscr",
                "value": 1.33,
                "unit": "x",
                "covenant_threshold": 1.20,
                "headroom": 0.13,
                "status": "OK",
            },
            {
                "metric_name": "llcr",
                "value": 1.55,
                "unit": "x",
                "covenant_threshold": 1.40,
                "headroom": 0.15,
                "status": "OK",
            },
        ]
    )


def _dummy_scenario_summary_df() -> pd.DataFrame:
    return pd.DataFrame(
        [
            {
                "scenario_name": "dutchbay_lendercase_2025Q4",
                "tag": "base",
                "project_irr": 0.125,
                "equity_irr": 0.165,
                "project_npv_lkr": 1_234_567_890.0,
                "equity_npv_lkr": 789_000_000.0,
                "dscr_min": 1.33,
                "llcr": 1.55,
                "plcr": 1.70,
                "capex_total_lkr": 30_000_000_000.0,
                "tariff_lkr_per_kwh": 20.30,
                "p50_generation_gwh": 650.0,
                "p90_generation_gwh": 580.0,
            }
        ]
    )


def test_executive_workbook_basic_structure(tmp_path) -> None:
    output_path = tmp_path / "executive_workbook_v1.xlsx"

    summary_df = _dummy_summary_df()
    cashflow_df = _dummy_cashflow_df()
    debt_df = _dummy_debt_df()
    ratios_df = _dummy_ratios_df()
    scenario_summary_df = _dummy_scenario_summary_df()

    result_path = build_executive_workbook(
        output_path=output_path,
        summary_df=summary_df,
        cashflow_df=cashflow_df,
        debt_df=debt_df,
        ratios_df=ratios_df,
        scenario_summary_df=scenario_summary_df,
    )

    assert isinstance(result_path, Path)
    assert result_path.exists()

    wb = load_workbook(result_path, read_only=True)
    sheet_names = set(wb.sheetnames)

    expected = {
        "Summary",
        "Cashflow",
        "DebtService",
        "Ratios",
        "ScenarioSummary",
    }

    # Workbook should contain at least these sheets
    missing = expected.difference(sheet_names)
    assert not missing, f"Missing expected sheets: {missing}"

    # Basic shape sanity checks
    summary_sheet = wb["Summary"]
    assert summary_sheet.max_row >= 3  # header + at least a couple of metrics
    assert summary_sheet.max_column >= 2  # section/metric/value style table
