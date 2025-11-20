#!/usr/bin/env python3
"""Smokes for analytics.export_helpers (Excel + charts).

These tests intentionally stay light-touch:
- ExcelExporter: verify sheets + formatting helpers don't crash and produce a valid workbook.
- ChartGenerator: verify each chart helper writes a PNG into the requested output_dir.
"""

from pathlib import Path

import pandas as pd

from analytics.export_helpers import ExcelExporter, ChartGenerator


def test_excel_exporter_end_to_end(tmp_path):
    """Basic E2E smoke for ExcelExporter.

    Creates a workbook with:
    - One DataFrame sheet
    - A conditional formatting rule
    - An embedded dummy chart image

    Asserts that the file exists and the expected sheet name is present.
    """
    # Arrange
    output_path = tmp_path / "export_helpers_smoke.xlsx"
    df = pd.DataFrame(
        {
            "scenario": ["base", "downside"],
            "npv": [123.45, 67.89],
        }
    )

    exporter = ExcelExporter(str(output_path))

    # Act: add data sheet
    exporter.add_dataframe_sheet(
        sheet_name="Summary",
        df=df,
        freeze_panes="B2",
        format_headers=True,
        auto_filter=True,
    )

    # Act: add a simple conditional formatting band on the NPV column
    exporter.add_conditional_formatting(
        sheet_name="Summary",
        column_range="B2:B3",
        rule_type="lessThan",
        threshold=100.0,
    )

    # Act: generate a tiny dummy chart image and embed it
    chart_path = tmp_path / "dummy_chart.png"
    cg = ChartGenerator(output_dir=str(tmp_path))
    cg.plot_npv_distribution(
        npv_values=[1_000_000.0, 2_000_000.0, 3_000_000.0],
        output_file=chart_path.name,
        bins=5,
    )

    exporter.add_chart_image(
        sheet_name="Summary",
        image_path=str(chart_path),
        cell="D2",
    )

    # Finally, save workbook
    exporter.save()

    # Assert: workbook exists and is non-empty
    assert output_path.exists(), "Expected Excel workbook to be created"
    assert output_path.stat().st_size > 0, "Excel workbook should not be empty"

    # Optional sanity: confirm the 'Summary' sheet exists
    from openpyxl import load_workbook

    wb = load_workbook(output_path)
    assert "Summary" in wb.sheetnames


def test_chart_generator_dscr_and_debt_smokes(tmp_path):
    """Smoke test DSCR and debt waterfall plotters.

    We only assert that PNG files are created in the requested directory.
    """
    cg = ChartGenerator(output_dir=str(tmp_path))

    dscr_data = {
        "base": [1.3, 1.4, 1.5],
        "downside": [1.1, 1.2, 1.3],
    }
    debt_data = {
        "base": [100e6, 90e6, 80e6],
        "downside": [100e6, 95e6, 90e6],
    }

    dscr_path = cg.plot_dscr_comparison(
        scenario_data=dscr_data,
        output_file="dscr_smoke.png",
        threshold=1.25,
    )
    debt_path = cg.plot_debt_waterfall(
        scenario_data=debt_data,
        output_file="debt_smoke.png",
    )

    assert Path(dscr_path).exists(), "DSCR chart PNG should be created"
    assert Path(debt_path).exists(), "Debt waterfall PNG should be created"


def test_chart_generator_kpi_and_distribution_smokes(tmp_path):
    """Smoke test KPI comparison + NPV distribution helpers."""
    cg = ChartGenerator(output_dir=str(tmp_path))

    kpi_df = pd.DataFrame(
        {
            "scenario": ["base", "downside", "upside"],
            "npv": [100e6, 60e6, 140e6],
        }
    )

    kpi_path = cg.plot_kpi_comparison(
        kpi_data=kpi_df,
        kpi_name="npv",
        output_file="npv_comparison_smoke.png",
    )

    npv_path = cg.plot_npv_distribution(
        npv_values=[50e6, 75e6, 100e6, 125e6, 150e6],
        output_file="npv_distribution_smoke.png",
        bins=10,
    )

    assert Path(kpi_path).exists(), "KPI comparison PNG should be created"
    assert Path(npv_path).exists(), "NPV distribution PNG should be created"
