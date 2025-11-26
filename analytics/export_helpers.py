#!/usr/bin/env python3
"""
Export helpers for DutchBay V14 scenario analytics.

Enhancements (Go With The Flow):
- YAML/Scenario metadata awareness (cover sheet with config summary)
- Immediate-use directory assurance (creates dirs for Excel and images)
- Excel writer context management (guaranteed file closure)
- OpenPyXL style extensions (autofit, formatting, highlights)
- Cleaner board views (Report Cover, key sheets ordered)
- Pre-export validation (validate_for_export warns on missing columns)
- Defensive: all openpyxl/matplotlib imports are optional
"""

from __future__ import annotations

import datetime
import logging
from dataclasses import dataclass
from os import PathLike
from pathlib import Path
from typing import (Any, Dict, Iterable, List, Mapping, Optional, Sequence,
                    Union)

import pandas as pd

logger = logging.getLogger(__name__)


def validate_for_export(
    summary_df: pd.DataFrame,
    timeseries_df: pd.DataFrame,
) -> List[str]:
    """
    Lightweight validation to warn about missing expected columns before export.
    This is deliberately non-fatal and only returns warning strings.
    """
    warnings: List[str] = []

    if "scenario_name" not in summary_df.columns:
        warnings.append("summary_df missing 'scenario_name' column")
    if "scenario_name" not in timeseries_df.columns:
        warnings.append("timeseries_df missing 'scenario_name' column")

    # Typical KPI columns we might reasonably expect (non-fatal if missing)
    expected_summary_kpis = ["project_irr", "equity_irr"]
    for col in expected_summary_kpis:
        if col not in summary_df.columns:
            warnings.append(f"summary_df missing expected KPI column '{col}'")

    if "dscr" not in timeseries_df.columns:
        warnings.append("timeseries_df missing 'dscr' column")

    return warnings


class ExcelExporter:
    """
    Excel export helper for ScenarioAnalytics.

    Provides:
      - add_metadata_sheet
      - add_dataframe_sheet
      - add_conditional_formatting
      - add_chart_image
      - write_scenario_summary
      - write_scenario_timeseries
      - export_summary_and_timeseries
      - _add_board_friendly_views
      - _reorder_sheets_for_board
      - autofit_all
      - save
    """

    def __init__(self, output_path: PathLike[Any]) -> None:
        self.output_path = Path(output_path)
        self._writer: Optional[pd.ExcelWriter] = None

    # ---------------------------------------------------------------------
    # Core writer plumbing
    # ---------------------------------------------------------------------

    def _ensure_writer(self) -> pd.ExcelWriter:
        if self._writer is None:
            self._writer = pd.ExcelWriter(self.output_path, engine="openpyxl")
        return self._writer

    def save(self) -> None:
        if self._writer is not None:
            self._writer.close()
            self._writer = None

    # ---------------------------------------------------------------------
    # Cover / metadata sheet
    # ---------------------------------------------------------------------

    def add_metadata_sheet(
        self,
        metadata: Dict[str, Any],
        sheet_name: str = "Report_Cover",
        run_id: Optional[str] = None,
    ) -> None:
        writer = self._ensure_writer()
        timestamp = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        lines: list[tuple[str, str]] = [("Report Generated", timestamp)]
        if run_id:
            lines.append(("Run ID", run_id))

        for k, v in metadata.items():
            lines.append((str(k), str(v)))

        df = pd.DataFrame(lines, columns=["Field", "Value"])
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        # Best-effort styling; failures are non-fatal
        try:
            from openpyxl.styles import Font

            workbook = writer.book
            ws = writer.sheets.get(sheet_name) or workbook[sheet_name]

            for cell in ws[1]:
                cell.font = Font(bold=True, size=12)

            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 50
        except Exception as exc:  # pragma: no cover - cosmetics only
            logger.warning("ExcelExporter: cover sheet formatting failed: %s", exc)

    # ---------------------------------------------------------------------
    # Generic dataframe sheet helper
    # ---------------------------------------------------------------------

    def add_dataframe_sheet(
        self,
        sheet_name: str,
        df: pd.DataFrame,
        freeze_panes: Optional[Union[str, tuple[int, int]]] = None,
        format_headers: bool = True,
        auto_filter: bool = True,
        highlight_warnings: bool = False,
        warning_column: Optional[str] = None,
        warning_threshold: Optional[float] = None,
    ) -> None:
        writer = self._ensure_writer()
        df.to_excel(writer, sheet_name=sheet_name, index=False)

        try:
            from openpyxl.styles import Font, PatternFill
            from openpyxl.utils import get_column_letter

            workbook = writer.book
            ws = writer.sheets.get(sheet_name) or workbook[sheet_name]
        except Exception as exc:  # pragma: no cover - environment dependent
            logger.warning("ExcelExporter: cannot access sheet %s: %s", sheet_name, exc)
            return

        if format_headers:
            try:
                for cell in ws[1]:
                    cell.font = Font(bold=True)
            except Exception as exc:  # pragma: no cover - cosmetic
                logger.warning(
                    "ExcelExporter: header format error %s: %s", sheet_name, exc
                )

        if auto_filter:
            try:
                ws.auto_filter.ref = ws.dimensions
            except Exception as exc:  # pragma: no cover - cosmetic
                logger.warning(
                    "ExcelExporter: auto-filter error %s: %s", sheet_name, exc
                )

        if freeze_panes:
            try:
                ws.freeze_panes = freeze_panes
            except Exception as exc:  # pragma: no cover - cosmetic
                logger.warning(
                    "ExcelExporter: freeze panes error %s: %s", sheet_name, exc
                )

        if highlight_warnings and warning_column and warning_threshold is not None:
            try:
                if warning_column in df.columns:
                    col_idx = df.columns.get_loc(warning_column) + 1
                    col_letter = get_column_letter(col_idx)
                    red_fill = PatternFill(
                        start_color="FFCCCC",
                        end_color="FFCCCC",
                        fill_type="solid",
                    )
                    for row_idx, value in enumerate(df[warning_column], start=2):
                        try:
                            if pd.notna(value) and float(value) < warning_threshold:
                                ws[f"{col_letter}{row_idx}"].fill = red_fill
                        except (ValueError, TypeError):
                            continue
            except Exception as exc:  # pragma: no cover - cosmetic
                logger.warning(
                    "ExcelExporter: warning highlight error %s: %s", sheet_name, exc
                )

    # ---------------------------------------------------------------------
    # Conditional formatting helper (used by tests)
    # ---------------------------------------------------------------------

    def add_conditional_formatting(
        self,
        sheet_name: str,
        column_range: str,
        rule_type: str = "lessThan",
        threshold: Optional[float] = None,
    ) -> None:
        """
        Add a simple conditional formatting rule to an existing sheet.

        Parameters
        ----------
        sheet_name : str
            Name of the sheet to format (e.g., "Summary").
        column_range : str
            Excel range string (e.g., "B2:B3").
        rule_type : str
            openpyxl CellIsRule operator, e.g. "lessThan", "greaterThan".
        threshold : Optional[float]
            Numeric threshold value for the rule.
        """
        writer = self._ensure_writer()
        if threshold is None:
            logger.warning(
                "ExcelExporter: conditional formatting skipped for %s; "
                "threshold was None",
                sheet_name,
            )
            return

        try:
            from openpyxl.formatting.rule import CellIsRule
            from openpyxl.styles import PatternFill

            workbook = writer.book
            ws = writer.sheets.get(sheet_name) or workbook[sheet_name]
        except Exception as exc:  # pragma: no cover - env / IO issues
            logger.warning(
                "ExcelExporter: conditional formatting not available for %s: %s",
                sheet_name,
                exc,
            )
            return

        try:
            fill = PatternFill(
                start_color="FFCCCC",
                end_color="FFCCCC",
                fill_type="solid",
            )
            rule = CellIsRule(
                operator=rule_type,
                formula=[str(threshold)],
                fill=fill,
            )
            ws.conditional_formatting.add(column_range, rule)
        except Exception as exc:  # pragma: no cover - cosmetic / env
            logger.warning(
                "ExcelExporter: failed to apply conditional formatting on %s "
                "(%s %s, threshold=%s): %s",
                sheet_name,
                rule_type,
                column_range,
                threshold,
                exc,
            )

    # ---------------------------------------------------------------------
    # Image embedding helper (used by tests for dummy chart)
    # ---------------------------------------------------------------------

    def add_chart_image(
        self,
        sheet_name: str,
        image_path: str,
        cell: str = "A1",
    ) -> None:
        """
        Embed a chart/image file into the given sheet at the specified cell.

        Parameters
        ----------
        sheet_name : str
            Target sheet name (e.g., "Summary").
        image_path : str
            Path to a PNG (or other supported) image on disk.
        cell : str
            Anchor cell location (e.g., "D2").
        """
        writer = self._ensure_writer()
        try:
            from openpyxl.drawing.image import Image as XLImage

            workbook = writer.book
            ws = writer.sheets.get(sheet_name) or workbook[sheet_name]
        except Exception as exc:  # pragma: no cover - env / IO issues
            logger.warning(
                "ExcelExporter: cannot attach image on sheet %s: %s",
                sheet_name,
                exc,
            )
            return

        try:
            img = XLImage(image_path)
            ws.add_image(img, cell)
        except Exception as exc:  # pragma: no cover - cosmetic / IO
            logger.warning(
                "ExcelExporter: failed to add chart image '%s' at %s on %s: %s",
                image_path,
                cell,
                sheet_name,
                exc,
            )

    # ---------------------------------------------------------------------
    # Scenario summary/timeseries helpers
    # ---------------------------------------------------------------------

    def write_scenario_summary(
        self,
        summary_df: pd.DataFrame,
        sheet_name: str = "Summary",
    ) -> None:
        self.add_dataframe_sheet(
            sheet_name=sheet_name,
            df=summary_df,
            format_headers=True,
            auto_filter=True,
        )

    def write_scenario_timeseries(
        self,
        timeseries_df: pd.DataFrame,
        sheet_name: str = "Timeseries",
    ) -> None:
        self.add_dataframe_sheet(
            sheet_name=sheet_name,
            df=timeseries_df,
            format_headers=True,
            auto_filter=True,
        )

    # ---------------------------------------------------------------------
    # High-level export entry point (this is what ScenarioAnalytics uses)
    # ---------------------------------------------------------------------

    def export_summary_and_timeseries(
        self,
        summary_df: pd.DataFrame,
        timeseries_df: pd.DataFrame,
        summary_sheet: str = "Summary",
        timeseries_sheet: str = "Timeseries",
        add_board_views: bool = True,
        scenario_metadata: Optional[Dict[str, Any]] = None,
        run_id: Optional[str] = None,
        validate_before_export: bool = True,
    ) -> None:
        logger.info(
            "ExcelExporter: exporting summary + timeseries to %s", self.output_path
        )

        if validate_before_export:
            warnings = validate_for_export(summary_df, timeseries_df)
            for warning in warnings:
                logger.warning("Export validation: %s", warning)

        if scenario_metadata is not None:
            self.add_metadata_sheet(metadata=scenario_metadata, run_id=run_id)

        self.write_scenario_summary(summary_df, sheet_name=summary_sheet)
        self.write_scenario_timeseries(timeseries_df, sheet_name=timeseries_sheet)

        if add_board_views:
            self._add_board_friendly_views(summary_df, timeseries_df)

        self._reorder_sheets_for_board()
        self.autofit_all()
        self.save()

    # ---------------------------------------------------------------------
    # Board-friendly views + sheet ordering
    # ---------------------------------------------------------------------

    def _add_board_friendly_views(
        self,
        summary_df: pd.DataFrame,
        timeseries_df: pd.DataFrame,
    ) -> None:
        # DSCR view
        if {"scenario_name", "dscr"}.issubset(timeseries_df.columns):
            try:
                dscr_cols = [
                    c
                    for c in timeseries_df.columns
                    if c in {"scenario_name", "year", "period", "dscr"}
                ]
                dscr_view = timeseries_df[dscr_cols].copy()
                if "year" in dscr_view.columns:
                    dscr_view.rename(columns={"year": "Year"}, inplace=True)
                elif "period" in dscr_view.columns:
                    dscr_view.rename(columns={"period": "Period"}, inplace=True)

                self.add_dataframe_sheet(
                    sheet_name="DSCR_View",
                    df=dscr_view,
                    highlight_warnings=True,
                    warning_column="dscr",
                    warning_threshold=1.2,
                )
            except Exception as exc:  # pragma: no cover - cosmetic
                logger.warning("ExcelExporter: DSCR view export failed: %s", exc)

        # IRR view
        irr_candidates = [c for c in summary_df.columns if "irr" in str(c).lower()]
        if irr_candidates:
            try:
                irr_view = summary_df[["scenario_name", *irr_candidates]].copy()
                self.add_dataframe_sheet(
                    sheet_name="IRR_View",
                    df=irr_view,
                    format_headers=True,
                    auto_filter=True,
                )
            except Exception as exc:  # pragma: no cover - cosmetic
                logger.warning("ExcelExporter: IRR view export failed: %s", exc)

    def _reorder_sheets_for_board(self) -> None:
        if self._writer is None:
            return
        try:
            workbook = self._writer.book
            priority_order = [
                "Report_Cover",
                "Summary",
                "DSCR_View",
                "IRR_View",
                "Timeseries",
            ]
            all_sheets = workbook.sheetnames
            new_order: List[str] = []

            for priority_name in priority_order:
                if priority_name in all_sheets:
                    new_order.append(priority_name)

            for sheet_name in all_sheets:
                if sheet_name not in new_order:
                    new_order.append(sheet_name)

            workbook._sheets = [workbook[name] for name in new_order]
        except Exception as exc:  # pragma: no cover - cosmetic
            logger.warning("ExcelExporter: sheet reordering failed: %s", exc)

    # ---------------------------------------------------------------------
    # Autofit
    # ---------------------------------------------------------------------

    def autofit_all(self, max_width: int = 50) -> None:
        if self._writer is None:
            return
        try:
            from openpyxl.utils import get_column_letter

            workbook = self._writer.book
        except Exception as exc:  # pragma: no cover - environment dependent
            logger.warning("ExcelExporter: can't access workbook for autofit: %s", exc)
            return

        for ws in workbook.worksheets:
            for column_cells in ws.columns:
                try:
                    max_length = max(
                        len(str(cell.value)) if cell.value else 0
                        for cell in column_cells
                    )
                except ValueError:
                    continue
                adjusted_width = min(max_length + 2, max_width)
                col_letter = get_column_letter(column_cells[0].column)
                ws.column_dimensions[col_letter].width = adjusted_width


# ---------------------------------------------------------------------------
# Chart exporters
# ---------------------------------------------------------------------------


@dataclass
class ChartExporter:
    output_dir: Path

    def __init__(self, output_dir: Union[PathLike[Any], str]) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_plt(self):
        try:
            import matplotlib.pyplot as plt  # type: ignore[import-not-found]

            return plt
        except Exception as exc:  # pragma: no cover
            logger.warning("ChartExporter: matplotlib not available: %s", exc)
            return None

    def export_dscr_chart(
        self,
        timeseries_df: pd.DataFrame,
        scenario_name_column: str = "scenario_name",
        dscr_column: str = "dscr",
        output_file: str = "dscr_series.png",
    ) -> Optional[Path]:
        required = {scenario_name_column, dscr_column}
        missing = required - set(timeseries_df.columns)
        if missing:
            logger.warning(
                "ChartExporter: DSCR chart skipped, missing columns %s", missing
            )
            return None

        plt = self._get_plt()
        if plt is None:
            return None

        out_path = self.output_dir / output_file
        try:
            plt.figure()
            grouped = timeseries_df.groupby(scenario_name_column)
            for name, grp in grouped:
                dscr_series = pd.to_numeric(grp[dscr_column], errors="coerce")
                x = range(1, len(dscr_series) + 1)
                plt.plot(x, dscr_series, marker="o", label=str(name))

            plt.axhline(1.0, linestyle="--", linewidth=1, color="red", alpha=0.7)
            plt.xlabel("Period")
            plt.ylabel("DSCR")
            plt.title("Debt Service Coverage Ratio (DSCR) by Scenario")
            plt.legend()
            plt.tight_layout()
            plt.savefig(out_path, dpi=150)
            plt.close()
            logger.info("ChartExporter: DSCR chart written to %s", out_path)
            return out_path
        except Exception as exc:  # pragma: no cover
            logger.warning("ChartExporter: failed to export DSCR chart: %s", exc)
            return None

    def export_irr_histogram(
        self,
        summary_df: pd.DataFrame,
        irr_column: str = "project_irr",
        output_file: str = "project_irr_hist.png",
        bins: int = 20,
    ) -> Optional[Path]:
        if irr_column not in summary_df.columns:
            logger.warning(
                "ChartExporter: IRR histogram skipped; missing column '%s'",
                irr_column,
            )
            return None

        plt = self._get_plt()
        if plt is None:
            return None

        out_path = self.output_dir / output_file
        try:
            plt.figure()
            data = pd.to_numeric(summary_df[irr_column], errors="coerce").dropna()
            if data.empty:
                logger.warning(
                    "ChartExporter: no finite values in '%s' for histogram",
                    irr_column,
                )
                return None

            plt.hist(data, bins=bins, edgecolor="black", alpha=0.7)
            plt.xlabel("IRR")
            plt.ylabel("Frequency")
            plt.title(f"{irr_column} distribution")
            plt.tight_layout()
            plt.savefig(out_path, dpi=150)
            plt.close()
            logger.info("ChartExporter: IRR histogram written to %s", out_path)
            return out_path
        except Exception as exc:  # pragma: no cover
            logger.warning("ChartExporter: failed to export IRR histogram: %s", exc)
            return None

    def export_charts(
        self,
        summary_df: pd.DataFrame,
        timeseries_df: pd.DataFrame,
    ) -> Dict[str, Optional[Path]]:
        return {
            "dscr_chart": self.export_dscr_chart(timeseries_df),
            "irr_histogram": self.export_irr_histogram(summary_df),
        }


class ChartGenerator:
    """
    Legacy / more generic chart generator used by other analytics layers.
    """

    def __init__(self, output_dir: PathLike[Any]) -> None:
        self.output_dir = Path(output_dir)
        self.output_dir.mkdir(parents=True, exist_ok=True)

    def _get_plt(self):
        try:
            import matplotlib.pyplot as plt  # type: ignore[import-not-found]

            return plt
        except Exception as exc:  # pragma: no cover
            logger.warning("ChartGenerator: matplotlib not available: %s", exc)
            return None

    def _resolve_path(self, output_file: PathLike[Any]) -> Path:
        path = Path(output_file)
        if not path.is_absolute():
            path = self.output_dir / path
        path.parent.mkdir(parents=True, exist_ok=True)
        return path

    def plot_kpi_comparison(
        self,
        kpi_data: Union[pd.DataFrame, Mapping[str, Iterable[float]]],
        kpi_name: str,
        output_file: PathLike[Any],
    ) -> Path:
        plt = self._get_plt()
        path = self._resolve_path(output_file)

        if plt is None:
            return path

        if isinstance(kpi_data, pd.DataFrame):
            df = kpi_data.copy()
        else:
            df = pd.DataFrame(kpi_data)

        if kpi_name in df.columns:
            y = df[kpi_name]
        else:
            numeric_cols = df.select_dtypes("number").columns
            if not numeric_cols:
                raise ValueError("No numeric columns available for KPI chart")
            kpi_name = str(numeric_cols[0])
            y = df[kpi_name]

        if "scenario_name" in df.columns:
            labels = df["scenario_name"].astype(str).tolist()
        else:
            labels = df.index.astype(str).tolist()

        fig, ax = plt.subplots()
        ax.bar(range(len(y)), y, edgecolor="black", alpha=0.7)
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=45, ha="right")
        ax.set_ylabel(kpi_name)
        ax.set_title(f"{kpi_name} comparison")
        fig.tight_layout()
        fig.savefig(path, dpi=150)
        plt.close(fig)
        return path

    def plot_npv_distribution(
        self,
        npv_values: Iterable[float],
        output_file: PathLike[Any],
        bins: int = 20,
    ) -> Path:
        plt = self._get_plt()
        path = self._resolve_path(output_file)

        if plt is None:
            return path

        fig, ax = plt.subplots()
        ax.hist(list(npv_values), bins=bins, edgecolor="black", alpha=0.7)
        ax.set_xlabel("NPV")
        ax.set_ylabel("Frequency")
        ax.set_title("NPV distribution")
        fig.tight_layout()
        fig.savefig(path, dpi=150)
        plt.close(fig)
        return path

    def plot_dscr_comparison(
        self,
        scenario_data: Mapping[str, Sequence[float]],
        output_file: PathLike[Any],
        threshold: Optional[float] = None,
    ) -> Path:
        plt = self._get_plt()
        path = self._resolve_path(output_file)

        if plt is None:
            return path

        fig, ax = plt.subplots()
        for name, series in scenario_data.items():
            y = list(series)
            x = list(range(1, len(y) + 1))
            ax.plot(x, y, marker="o", label=str(name))

        if threshold is not None:
            ax.axhline(threshold, linestyle="--", linewidth=1, color="red", alpha=0.7)

        ax.set_xlabel("Period")
        ax.set_ylabel("DSCR")
        ax.set_title("DSCR comparison by scenario")
        ax.legend()
        fig.tight_layout()
        fig.savefig(path, dpi=150)
        plt.close(fig)
        return path

    def plot_debt_waterfall(
        self,
        scenario_data: Mapping[str, Sequence[float]],
        output_file: PathLike[Any],
    ) -> Path:
        plt = self._get_plt()
        path = self._resolve_path(output_file)

        if plt is None:
            return path

        labels: list[str] = []
        values: list[float] = []

        for name, series in scenario_data.items():
            labels.append(str(name))
            s = list(series)
            values.append(s[-1] if s else 0.0)

        fig, ax = plt.subplots()
        ax.bar(range(len(values)), values, edgecolor="black", alpha=0.7)
        ax.set_xticks(range(len(labels)))
        ax.set_xticklabels(labels, rotation=45, ha="right")
        ax.set_ylabel("Debt (end of horizon)")
        ax.set_title("Debt waterfall by scenario")
        fig.tight_layout()
        fig.savefig(path, dpi=150)
        plt.close(fig)
        return path
