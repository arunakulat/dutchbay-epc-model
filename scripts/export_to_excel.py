#!/usr/bin/env python3
"""
Excel export script for DutchBay EPC Model analysis results.

Exports complete analysis to Excel with multiple sheets:
- Annual Cashflow: P&L with debt service components
- Debt Service Detail: Interest vs Principal breakdown
- Debt Covenants: DSCR and tenor tracking
- KPI Summary: Financial metrics
- Optional: Refinancing if calculated
"""
import json
import logging
import sys
from pathlib import Path
from typing import Any

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl required. Install with: pip install openpyxl")
    sys.exit(1)

from analytics.pipeline_v14 import run_v14_pipeline

# Configure logging
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s - %(name)s - %(levelname)s - %(message)s",
)
logger = logging.getLogger(__name__)

# Styles for Excel
HEADER_FILL = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
HEADER_FONT = Font(bold=True, color="FFFFFF", size=11)
SUBHEADER_FILL = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
SUBHEADER_FONT = Font(bold=True, size=10)
TOTAL_FILL = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
TOTAL_FONT = Font(bold=True, size=10)
BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)


def add_annual_cashflow_sheet(wb: openpyxl.Workbook, results: dict[str, Any]) -> None:
    """Add Annual Cashflow sheet with complete P&L and debt service."""
    ws = wb.create_sheet("Annual Cashflow", 0)
    
    annual_rows = results.get('annual_rows', [])
    if not annual_rows:
        logger.warning("No annual_rows found in results")
        return
    
    # Headers
    headers = [
        "Year", "Revenue (LKR)", "EBITDA (LKR)", "CFADS (LKR)",
        "Debt Service (USD)", "Interest (USD)", "Principal (USD)", "DSCR"
    ]
    
    ws.append(headers)
    header_row = ws[1]
    for cell in header_row:
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_data in annual_rows:
        year = int(row_data.get('year', 0))
        revenue = float(row_data.get('revenue_lkr', 0))
        ebitda = float(row_data.get('ebitda_lkr', 0))
        cfads = float(row_data.get('cfads_final_lkr', 0))
        debt_svc = float(row_data.get('debt_service_usd', 0))
        interest = float(row_data.get('interest_usd', 0))
        principal = float(row_data.get('principal_repayment_usd', 0))
        dscr = float(row_data.get('dscr', 0))
        
        ws.append([
            year,
            revenue,
            ebitda,
            cfads,
            debt_svc,
            interest,
            principal,
            dscr if dscr > 0 else None
        ])
    
    # Format columns
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F', 'G']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['H'].width = 10
    
    # Format cells
    for row_idx in range(2, len(annual_rows) + 2):
        ws[f'A{row_idx}'].alignment = Alignment(horizontal="center")
        for col in ['B', 'C', 'D', 'E', 'F', 'G']:
            cell = ws[f'{col}{row_idx}']
            cell.number_format = '#,##0'
            cell.alignment = Alignment(horizontal="right")
            cell.border = BORDER
        ws[f'H{row_idx}'].number_format = '0.00'
        ws[f'H{row_idx}'].alignment = Alignment(horizontal="center")
        ws[f'H{row_idx}'].border = BORDER
    
    logger.info("Added Annual Cashflow sheet with %d rows", len(annual_rows))


def add_debt_service_detail_sheet(wb: openpyxl.Workbook, results: dict[str, Any]) -> None:
    """Add Debt Service Detail sheet focusing on interest breakdown."""
    ws = wb.create_sheet("Debt Service Detail", 1)
    
    annual_rows = results.get('annual_rows', [])
    if not annual_rows:
        return
    
    # Title
    ws['A1'] = "DEBT SERVICE ANALYSIS - Interest vs Principal Breakdown"
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:H1')
    
    # Headers
    headers = [
        "Year", "Debt Service (USD)", "Interest (USD)", "Interest %",
        "Principal (USD)", "Principal %", "DSCR", "Status"
    ]
    ws.append([])  # Blank row
    ws.append(headers)
    
    header_row = ws[3]
    for cell in header_row:
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = BORDER
        cell.alignment = Alignment(horizontal="center", vertical="center")
    
    # Data rows
    for row_data in annual_rows:
        year = int(row_data.get('year', 0))
        debt_svc = float(row_data.get('debt_service_usd', 0))
        interest = float(row_data.get('interest_usd', 0))
        principal = float(row_data.get('principal_repayment_usd', 0))
        dscr = float(row_data.get('dscr', 0))
        
        interest_pct = interest / debt_svc if debt_svc > 0 else 0
        principal_pct = principal / debt_svc if debt_svc > 0 else 0
        
        status = "Interest-Only" if principal == 0 else ("Amortizing" if dscr > 0 else "Grace")
        
        ws.append([
            year,
            debt_svc,
            interest,
            interest_pct,
            principal,
            principal_pct,
            dscr if dscr > 0 else None,
            status
        ])
    
    # Format columns
    ws.column_dimensions['A'].width = 8
    for col in ['B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 16
    ws.column_dimensions['G'].width = 10
    ws.column_dimensions['H'].width = 15
    
    # Format cells
    for row_idx in range(4, len(annual_rows) + 4):
        ws[f'A{row_idx}'].alignment = Alignment(horizontal="center")
        ws[f'B{row_idx}'].number_format = '"$"#,##0'
        ws[f'C{row_idx}'].number_format = '"$"#,##0'
        ws[f'D{row_idx}'].number_format = '0.0%'
        ws[f'E{row_idx}'].number_format = '"$"#,##0'
        ws[f'F{row_idx}'].number_format = '0.0%'
        ws[f'G{row_idx}'].number_format = '0.00'
        ws[f'H{row_idx}'].alignment = Alignment(horizontal="left")
        
        for col in ['A', 'B', 'C', 'D', 'E', 'F', 'G', 'H']:
            ws[f'{col}{row_idx}'].border = BORDER
    
    logger.info("Added Debt Service Detail sheet")


def add_debt_covenants_sheet(wb: openpyxl.Workbook, results: dict[str, Any]) -> None:
    """Add Debt Covenants sheet tracking DSCR and tenor."""
    ws = wb.create_sheet("Debt Covenants", 2)
    
    annual_rows = results.get('annual_rows', [])
    debt_result = results.get('debt_result', {})
    
    if not annual_rows:
        return
    
    # Summary stats
    ws['A1'] = "DEBT COVENANT COMPLIANCE"
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:D1')
    
    min_dscr = float(debt_result.get('min_dscr', 0))
    tenor_years = int(debt_result.get('tenor_years', 15))
    construction_years = int(debt_result.get('construction_years', 2))
    
    ws['A3'] = "Metric"
    ws['B3'] = "Value"
    for cell in [ws['A3'], ws['B3']]:
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = BORDER
    
    ws['A4'] = "Minimum DSCR"
    ws['B4'] = min_dscr
    ws['B4'].number_format = '0.00'
    
    ws['A5'] = "Tenor (years)"
    ws['B5'] = tenor_years
    
    ws['A6'] = "Construction Period (years)"
    ws['B6'] = construction_years
    
    # DSCR detail
    ws['A9'] = "Year-by-Year DSCR Analysis"
    ws['A9'].font = HEADER_FONT
    ws['A9'].fill = HEADER_FILL
    ws.merge_cells('A9:C9')
    
    headers = ["Year", "DSCR", "Target", "Status"]
    ws.append([])
    ws.append(headers)
    
    header_row = ws[11]
    for cell in header_row:
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = BORDER
    
    target_dscr = 1.30
    
    for row_data in annual_rows:
        year = int(row_data.get('year', 0))
        dscr = float(row_data.get('dscr', 0))
        status = "✓ PASS" if dscr >= target_dscr or dscr == 0 else "✗ FAIL"
        
        ws.append([year, dscr if dscr > 0 else None, target_dscr, status])
    
    # Format
    ws.column_dimensions['A'].width = 12
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 12
    ws.column_dimensions['D'].width = 12
    
    for row_idx in range(12, len(annual_rows) + 12):
        ws[f'B{row_idx}'].number_format = '0.00'
        ws[f'C{row_idx}'].number_format = '0.00'
        for col in ['A', 'B', 'C', 'D']:
            ws[f'{col}{row_idx}'].border = BORDER
    
    logger.info("Added Debt Covenants sheet")


def add_kpi_summary_sheet(wb: openpyxl.Workbook, results: dict[str, Any]) -> None:
    """Add KPI Summary sheet with financial metrics."""
    ws = wb.create_sheet("KPI Summary", 3)
    
    kpis = results.get('kpis', {})
    debt_result = results.get('debt_result', {})
    
    ws['A1'] = "KEY PERFORMANCE INDICATORS"
    ws['A1'].font = HEADER_FONT
    ws['A1'].fill = HEADER_FILL
    ws.merge_cells('A1:C1')
    
    ws['A3'] = "Metric"
    ws['B3'] = "Value"
    ws['C3'] = "Unit"
    for cell in [ws['A3'], ws['B3'], ws['C3']]:
        cell.fill = SUBHEADER_FILL
        cell.font = SUBHEADER_FONT
        cell.border = BORDER
    
    row = 4
    total_interest = sum(float(r.get('interest_usd', 0)) for r in results.get('annual_rows', []))
    
    metrics = [
        ("Project NPV", kpis.get('project_npv', 0), "USD"),
        ("Project IRR", kpis.get('project_irr', 0), "% p.a."),
        ("Min DSCR", debt_result.get('min_dscr', 0), "ratio"),
        ("Max Debt", kpis.get('max_debt_usd', 0), "USD"),
        ("Total Interest Paid", total_interest, "USD"),
    ]
    
    for metric_name, value, unit in metrics:
        ws[f'A{row}'] = metric_name
        ws[f'B{row}'] = value
        ws[f'C{row}'] = unit
        
        ws[f'A{row}'].border = BORDER
        ws[f'B{row}'].border = BORDER
        ws[f'C{row}'].border = BORDER
        
        if isinstance(value, (int, float)):
            if "USD" in unit or "Debt" in metric_name or "Interest" in metric_name:
                ws[f'B{row}'].number_format = '"$"#,##0'
            elif "%" in unit:
                ws[f'B{row}'].number_format = '0.00%'
            else:
                ws[f'B{row}'].number_format = '0.00'
        
        row += 1
    
    ws.column_dimensions['A'].width = 25
    ws.column_dimensions['B'].width = 18
    ws.column_dimensions['C'].width = 12
    
    logger.info("Added KPI Summary sheet")


def export_to_excel(config_path: str, output_path: str) -> None:
    """Run complete analysis and export to Excel.
    
    Parameters
    ----------
    config_path : str
        Path to scenario YAML config
    output_path : str
        Path for output Excel file
    """
    logger.info("Starting export: config=%s, output=%s", config_path, output_path)
    
    # Run pipeline
    logger.info("Running pipeline...")
    results = run_v14_pipeline(
        config_path,
        validation_mode='off',
    )
    
    # Create workbook
    wb = openpyxl.Workbook()
    wb.remove(wb.active)  # Remove default sheet
    
    # Add sheets
    add_annual_cashflow_sheet(wb, results)
    add_debt_service_detail_sheet(wb, results)
    add_debt_covenants_sheet(wb, results)
    add_kpi_summary_sheet(wb, results)
    
    # Save workbook
    output_file = Path(output_path)
    output_file.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_file)
    
    logger.info("✅ Excel export complete: %s", output_file)
    print(f"\n✅ Analysis exported to: {output_file}")
    print(f"   - Annual Cashflow sheet with full P&L")
    print(f"   - Debt Service Detail with interest breakdown")
    print(f"   - Debt Covenants tracking")
    print(f"   - KPI Summary")


if __name__ == '__main__':
    if len(sys.argv) < 3:
        print("Usage: python export_to_excel.py <config_yaml> <output_xlsx>")
        print("\nExample:")
        print("  python export_to_excel.py scenarios/dutchbay_lendercase_2025Q4.yaml out/analysis.xlsx")
        sys.exit(1)
    
    config_path = sys.argv[1]
    output_path = sys.argv[2]
    
    try:
        export_to_excel(config_path, output_path)
    except Exception as e:
        logger.error("Export failed: %s", e, exc_info=True)
        sys.exit(1)
